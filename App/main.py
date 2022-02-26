import openpyxl
import math
PATHS = {
    'australia': 'AU.xlsx',
    'brazil': 'br.xlsx',
    'egypt': 'EG.xlsx'
}

def display_population(sheet) : 
    row = 1
    total = 0
    while True:
        city = sheet.cell(row = row, column=1).value
        if not(city):
            break
        population = int(sheet.cell(row = row, column=2).value)
        total += population
        print(f'{city} : {population}')
        row += 1
    print(f'Total population is : {total}')

        
def max_min(sheet) : 
    mx = 0
    mn = math.inf
    mx_city = ''
    mn_city = ''
    row = 1
    while True:
        city = sheet.cell(row = row, column=1).value
        if not(city):
            break
        population = int(sheet.cell(row = row, column=2).value)
        if population > mx:
            mx = population
            mx_city = city
        if population < mn :
            mn = population
            mn_city = city

        row += 1
    print(f'Highest Population City is {mx_city} with {mx} citizen')
    print(f'lowest Population City is {mn_city} with {mn} citizen')
def get_country():
    country = input('Please Choose Country From The Countries Available :  ').lower()
    while country not in PATHS:
        country = input('Please Choose Country From The Countries Available :  ').lower()
    return country

def main():
    print(','.join(PATHS.keys()) + ' are the available countries')
    country = get_country()
    print("""
commands :-
    1.N : choose another country
    2.P :  Display the population of each state / province / governorate and total population of the country,
    3.M :  Display the state / province / governorate with the highest population and the one with the lowest population,
    4.Q : exit
    """)
    command = input('please enter your command : ').lower()
    command_funtion_map = {
        'p' : display_population,
        'm' : max_min
    }
    while command != 'q':
        if command == 'n':
            country = get_country()
            command = input('please enter your command : ').lower()
            continue
        path = PATHS[country]
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.active
        sheet_obj.iter_rows
        command_funtion_map[command](sheet_obj)
        command = input('please enter your command : ').lower()
    


if __name__ == '__main__':
    main()
